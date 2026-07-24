"""
Convert US News lookup final CSV to formatted Excel workbook.
Color-coded ratings, auto-width columns, frozen header, summary sheet.

Usage:  python excel-export-script.py

Input:  usnews_lookup_final.csv  (pipe-delimited: row|name|city|state|rating)
Output: US News Nursing Home Ratings 2026.xlsx

Requires: pip install openpyxl
"""

import csv
from collections import Counter

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

INPUT = 'usnews_lookup_final.csv'
OUTPUT = 'US News Nursing Home Ratings 2026.xlsx'

# Read CSV
rows = []
with open(INPUT, 'r') as f:
    reader = csv.DictReader(f, delimiter='|')
    fieldnames = reader.fieldnames
    rows = list(reader)

print(f"Loaded {len(rows)} facilities from {INPUT}")

wb = openpyxl.Workbook()

# ── Sheet 1: Data ───────────────────────────────────────────
ws = wb.active
ws.title = "US News Ratings 2026"

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hp_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
nf_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ae_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

for ci, field in enumerate(fieldnames, 1):
    cell = ws.cell(row=1, column=ci, value=field)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

for ri, row in enumerate(rows, 2):
    for ci, field in enumerate(fieldnames, 1):
        cell = ws.cell(row=ri, column=ci, value=row[field])
    rating_cell = ws.cell(row=ri, column=fieldnames.index('rating') + 1)
    if 'High Performing' in row['rating']:
        rating_cell.fill = hp_fill
    elif 'NOT FOUND' in row['rating']:
        rating_cell.fill = nf_fill
    elif 'As Expected' in row['rating']:
        rating_cell.fill = ae_fill

# Auto-width
for col in ws.columns:
    max_len = max((len(str(c.value or '')) for c in col), default=0)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

ws.freeze_panes = 'A2'

# ── Sheet 2: Summary ────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
ws2['A1'] = "US News Nursing Home Ratings Summary"
ws2['A1'].font = Font(bold=True, size=14)

for ci, h in enumerate(['Rating', 'Count', 'Percentage'], 1):
    c = ws2.cell(row=3, column=ci, value=h)
    c.font = header_font
    c.fill = header_fill

cnt = Counter(r['rating'] for r in rows)
total = len(rows)
for i, (rating, n) in enumerate(sorted(cnt.items(), key=lambda x: -x[1]), 4):
    ws2.cell(row=i, column=1, value=rating)
    ws2.cell(row=i, column=2, value=n)
    ws2.cell(row=i, column=3, value=f"{n/total*100:.1f}%")

ws2.column_dimensions['A'].width = 55
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 12

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
