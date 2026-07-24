#!/usr/bin/env python3
"""Verify spreadsheet CCNs against live CMS Provider Data.

Usage:
    python3 verify-ccns.py <input.xlsx> [--sheet "Sheet Name"] [--ccn-col "CMS CCN"] [--field-defs FIELDS.json]

Input: xlsx with a "CMS CCN" column containing CMS Certification Numbers
Output: side-by-side comparison of spreadsheet values vs live CMS API data

Field names are auto-detected from CMS CSV headers. Pass --field-defs to override
the mapping from spreadsheet column names to CMS API field names.
"""
import argparse
import csv
import json
import sys
import urllib.request
from collections import OrderedDict

# Default CSV download URL (check data.cms.gov for latest)
DEFAULT_CMS_URL = (
    "https://data.cms.gov/provider-data/sites/default/files/resources/"
    "bc7015f6a981fa7e209809e021f8f0cc_1781194538/NH_ProviderInfo_Jun2026.csv"
)

# Default field mappings: spreadsheet_column -> CMS_csv_header
DEFAULT_FIELDS = OrderedDict([
    ("Overall Rating", "Overall Rating"),
    ("Staffing Rating", "Staffing Rating"),
    ("Health Inspection Rating", "Health Inspection Rating"),
    ("QM Rating", "QM Rating"),
    ("CNA Hrs/Res/Day", "Reported Nurse Aide Staffing Hours per Resident per Day"),
    ("RN Hrs/Res/Day", "Reported RN Staffing Hours per Resident per Day"),
    ("Total Nurse Hrs/Res/Day", "Reported Total Nurse Staffing Hours per Resident per Day"),
    ("Total Nurse Turnover %", "Total nursing staff turnover"),
    ("RN Turnover %", "Registered Nurse turnover"),
    ("# Fines", "Number of Fines"),
    ("Fine Amount ($)", "Total Amount of Fines in Dollars"),
    ("# Payment Denials", "Number of Payment Denials"),
    ("Total Penalties", "Total Number of Penalties"),
    ("Certified Beds", "Number of Certified Beds"),
    ("Avg Residents/Day", "Average Number of Residents per Day"),
])


def download_cms_csv(url=None):
    """Download and parse the full CMS provider CSV. Returns dict keyed by CCN."""
    url = url or DEFAULT_CMS_URL
    print(f"Downloading CMS data from {url}...", file=sys.stderr)
    
    providers = {}
    with urllib.request.urlopen(url, timeout=120) as resp:
        reader = csv.DictReader(r.decode('utf-8') for r in resp)
        for row in reader:
            ccn = row.get("CMS Certification Number (CCN)", "").strip()
            if ccn:
                providers[ccn] = row
    
    print(f"  Loaded {len(providers)} providers (processing_date={providers.get(list(providers.keys())[0], {}).get('Processing Date', 'unknown') if providers else 'N/A'})", file=sys.stderr)
    return providers


def compare_facility(ccn, spreadsheet_row, cms_data, fields):
    """Compare one facility's spreadsheet values against CMS data."""
    cms_row = cms_data.get(ccn)
    if not cms_row:
        return {"ccn": ccn, "found": False, "comparisons": []}
    
    comparisons = []
    for ss_key, cms_key in fields.items():
        ss_val = str(spreadsheet_row.get(ss_key, "")).strip()
        cms_val = str(cms_row.get(cms_key, "")).strip()
        
        # Normalize for numeric comparison
        def norm_num(v):
            try:
                return f"{float(v):.10f}".rstrip('0').rstrip('.')
            except (ValueError, TypeError):
                return v.lower().strip()
        
        match = norm_num(ss_val) == norm_num(cms_val)
        comparisons.append({
            "metric": ss_key,
            "spreadsheet": ss_val,
            "cms_api": cms_val,
            "match": match,
        })
    
    return {
        "ccn": ccn,
        "found": True,
        "cms_name": cms_row.get("Provider Name", ""),
        "cms_city": cms_row.get("City/Town", ""),
        "cms_state": cms_row.get("State", ""),
        "comparisons": comparisons,
    }


def format_report(results, fields):
    """Generate a formatted verification report."""
    lines = []
    sep = "=" * 130
    sub = "-" * 130
    
    total_comps = 0
    total_matches = 0
    total_mismatches = 0
    facilities_with_mismatches = 0
    
    for r in results:
        lines.append(sep)
        if not r["found"]:
            lines.append(f"❌ CCN {r['ccn']} - NOT FOUND in CMS dataset")
            continue
        
        lines.append(f"✅ CCN {r['ccn']} - {r['cms_name']} ({r['cms_city']}, {r['cms_state']})")
        lines.append(sep)
        
        header = f"{'Metric':<35} {'Spreadsheet':<20} {'CMS API':<20} {'Status'}"
        lines.append(header)
        lines.append(sub)
        
        has_mismatch = False
        for c in r["comparisons"]:
            total_comps += 1
            if c["match"]:
                total_matches += 1
                status = "✅ MATCH"
            else:
                total_mismatches += 1
                has_mismatch = True
                status = "❌ MISMATCH"
            lines.append(f"{c['metric']:<35} {c['spreadsheet']:<20} {c['cms_api']:<20} {status}")
        
        if has_mismatch:
            facilities_with_mismatches += 1
        lines.append("")
    
    # Summary
    lines.append("\n" + "=" * 130)
    lines.append("SUMMARY")
    lines.append("=" * 130)
    lines.append(f"Facilities checked: {len(results)}")
    lines.append(f"Total field comparisons: {total_comps}")
    lines.append(f"Matches: {total_matches}")
    lines.append(f"Mismatches: {total_mismatches}")
    lines.append(f"Facilities with at least one mismatch: {facilities_with_mismatches}")
    
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="Verify CCNs against live CMS data")
    parser.add_argument("xlsx", nargs="?", help="Input spreadsheet path")
    parser.add_argument("--sheet", default="Sheet1", help="Sheet name")
    parser.add_argument("--ccn-col", default="CMS CCN", help="Column name for CCNs")
    parser.add_argument("--cms-url", help="CMS CSV URL (auto-downloads if not specified)")
    parser.add_argument("--ccn-list", nargs="+", help="Inline CCN list (skip xlsx)")
    parser.add_argument("--field-defs", help="JSON file mapping spreadsheet cols to CMS headers")
    parser.add_argument("--output", help="Save report to file")
    args = parser.parse_args()
    
    # Load field definitions
    fields = DEFAULT_FIELDS
    if args.field_defs:
        with open(args.field_defs) as f:
            fields = OrderedDict(json.load(f))
    
    # Get CCNs and spreadsheet data
    ccns_to_check = []
    spreadsheet_rows = {}
    
    if args.ccn_list:
        for ccn in args.ccn_list:
            ccns_to_check.append(ccn)
            spreadsheet_rows[ccn] = {}
    elif args.xlsx:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(args.xlsx, data_only=True)
            ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb.active
            
            headers = [cell.value for cell in ws[1]]
            ccn_idx = None
            field_indices = {}
            for i, h in enumerate(headers):
                if h and str(h).strip() == args.ccn_col:
                    ccn_idx = i
                for ss_key in fields:
                    if h and str(h).strip() == ss_key:
                        field_indices[ss_key] = i
            
            if ccn_idx is None:
                print(f"Error: Column '{args.ccn_col}' not found in spreadsheet", file=sys.stderr)
                sys.exit(1)
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                ccn = str(row[ccn_idx]).strip() if row[ccn_idx] else ""
                if ccn and ccn not in ('None', ''):
                    ccns_to_check.append(ccn)
                    spreadsheet_rows[ccn] = {}
                    for ss_key, idx in field_indices.items():
                        if idx is not None and idx < len(row):
                            val = row[idx]
                            spreadsheet_rows[ccn][ss_key] = val if val is not None else ""
        except ImportError:
            print("Error: openpyxl required for xlsx input. pip install openpyxl", file=sys.stderr)
            sys.exit(1)
    else:
        print("Error: provide either --ccn-list or a spreadsheet path", file=sys.stderr)
        sys.exit(1)
    
    # Download CMS data
    cms_data = download_cms_csv(args.cms_url)
    
    # Compare
    results = []
    for ccn in ccns_to_check:
        result = compare_facility(ccn, spreadsheet_rows.get(ccn, {}), cms_data, fields)
        results.append(result)
    
    # Generate report
    report = format_report(results, fields)
    
    if args.output:
        with open(args.output, 'w') as f:
            f.write(report)
        print(f"Report saved to {args.output}", file=sys.stderr)
    else:
        print(report)


if __name__ == "__main__":
    main()
