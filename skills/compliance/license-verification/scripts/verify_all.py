#!/usr/bin/env python
"""
Master license verifier for all states.
Reads Excel, runs each state's scraper, writes results to output Excel.

Usage:
    python verify_all.py

Output:
    D:/license-verification/results/verification_YYYY-MM-DD.xlsx
"""
import time
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Adjust path if running from skill dir vs project dir
import sys
sys.path.insert(0, str(Path(__file__).parent.parent.parent.parent / "D:/license-verification"))

from states import (
    alabama, alaska, arizona, california, colorado,
    idaho, iowa, kansas, nebraska, nevada,
    oregon, south_carolina, tennessee, texas, utah,
    washington, wisconsin,
)

EXCEL_PATH = "C:/Users/kevin/Desktop/ENSG Facilities Only 6.1.26.xlsx"
RESULTS_DIR = Path("D:/license-verification/results")
RESULTS_DIR.mkdir(exist_ok=True)

# Map state name -> verify function
STATE_VERIFIERS = {
    "ALABAMA": alabama.verify_alabama,
    "ALASKA": alaska.verify_alaska,
    "ARIZONA": arizona.verify_arizona,
    "CALIFORNIA": california.verify_california,
    "COLORADO": colorado.verify_colorado,
    "IDAHO": idaho.verify_idaho,
    "IOWA": iowa.verify_iowa,
    "KANSAS": kansas.verify_kansas,
    "NEBRASKA": nebraska.verify_nebraska,
    "NEVADA": nevada.verify_nevada,
    "OREGON": oregon.verify_oregon,
    "SOUTH CAROLINA": south_carolina.verify_south_carolina,
    "TENNESSEE": tennessee.verify_tennessee,
    "TEXAS": texas.verify_texas,
    "UTAH": utah.verify_utah,
    "WASHINGTON": washington.verify_washington,
    "WISCONSIN": wisconsin.verify_wisconsin,
}

# Status mapping to PASS/FAIL/NEEDS MANUAL REVIEW
PASS_STATUSES = {"ACTIVE", "FOUND", "Active", "Actively Licensed", "Current"}
FAIL_STATUSES = {"NOT FOUND", "INACTIVE", "EXPIRED", "REVOKED", "SUSPENDED", "DENIED",
                 "Lapsed", "Expired", "Inactive", "Revoked", "Suspended", "Denied"}
MANUAL_STATUSES = {"BLOCKED", "NEEDS_MANUAL", "ERROR", "Blocked", "Error"}


def classify_status(status):
    if not status:
        return "NEEDS MANUAL REVIEW"
    s = status.upper()
    if s in PASS_STATUSES:
        return "PASS"
    if s in FAIL_STATUSES:
        return "FAIL"
    return "NEEDS MANUAL REVIEW"


def load_excel():
    """Load admins from Excel, grouped by state."""
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    facilities = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        admin = str(row[5]).strip() if row[5] else ""
        state = str(row[8]).strip().upper() if row[8] else ""
        facility = str(row[0]).strip() if row[0] else ""
        if state and admin:
            facilities.append({
                "state": state,
                "admin": admin,
                "facility": facility,
            })
    return facilities


def run_verification():
    """Run all verifications and write results."""
    facilities = load_excel()
    
    by_state = {}
    no_admin = []
    for fac in facilities:
        if fac["state"] in STATE_VERIFIERS:
            by_state.setdefault(fac["state"], []).append(fac)
        else:
            no_admin.append(fac)

    results = []
    errors = []

    print(f"Loaded {len(facilities)} facilities with admins")
    print(f"States to verify: {len(by_state)}")
    print()

    for state in sorted(by_state.keys()):
        verifier = STATE_VERIFIERS[state]
        state_facilities = by_state[state]
        
        print(f"[{state}] Verifying {len(state_facilities)} admins...")
        start = time.time()
        
        for fac in state_facilities:
            admin = fac["admin"]
            try:
                result = verifier(admin)
            except Exception as e:
                result = {
                    "status": "ERROR",
                    "expiration": "",
                    "url": "",
                    "note": str(e),
                    "days_until_expiry": None,
                }
            
            overall = classify_status(result.get("status", ""))
            days_until = result.get("days_until_expiry")
            expiration = result.get("expiration", "")
            alert = ""
            if days_until is not None and days_until >= 0 and days_until <= 60:
                alert = f"EXPIRES IN {days_until} DAYS"
            elif days_until is not None and days_until < 0:
                alert = "EXPIRED"

            results.append({
                "state": state,
                "facility": fac["facility"],
                "admin": admin,
                "status": result.get("status", ""),
                "overall": overall,
                "alert": alert,
                "expiration": expiration,
                "note": result.get("note", ""),
                "days_until_expiry": days_until if days_until is not None else "",
            })
        
        elapsed = time.time() - start
        print(f"  Done in {elapsed:.1f}s")

    output_path = RESULTS_DIR / f"verification_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    write_results_excel(results, no_admin, errors, output_path)
    print(f"\nResults written to: {output_path}")
    return output_path


def write_results_excel(results, no_admin, errors, path):
    """Write results to a formatted Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Verification Results"

    headers = ["State", "Facility", "Admin Name", "Status", "Result",
               "Expiration Alert", "Expiration Date", "Days Until Expiry", "URL", "Note"]
    ws.append(headers)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_font = Font(color="006100")
    red_font = Font(color="9C0006")
    dark_font = Font(color="9C5700")

    for row in results:
        result_val = row["overall"]
        alert_val = row["alert"]
        
        row_data = [
            row["state"],
            row["facility"],
            row["admin"],
            row["status"],
            result_val,
            alert_val,
            row["expiration"],
            row["days_until_expiry"] if row["days_until_expiry"] != "" else "",
            row.get("url", ""),
            row["note"],
        ]
        ws.append(row_data)

        result_cell = ws.cell(row=ws.max_row, column=5)
        if result_val == "PASS":
            result_cell.fill = green_fill
            result_cell.font = green_font
        elif result_val == "FAIL":
            result_cell.fill = red_fill
            result_cell.font = red_font
        else:
            result_cell.fill = yellow_fill
            result_cell.font = dark_font

        if alert_val:
            alert_cell = ws.cell(row=ws.max_row, column=6)
            alert_cell.font = Font(bold=True, color="9C0006")

    # Summary
    summary_start = ws.max_row + 2
    ws.cell(row=summary_start, column=1, value="SUMMARY")
    ws.cell(row=summary_start, column=1).font = Font(bold=True)

    total = len(results)
    passes = sum(1 for r in results if r["overall"] == "PASS")
    fails = sum(1 for r in results if r["overall"] == "FAIL")
    manual = sum(1 for r in results if r["overall"] == "NEEDS MANUAL REVIEW")
    alerts = sum(1 for r in results if r["alert"])

    summary = [
        f"Total facilities checked: {total}",
        f"PASS: {passes}",
        f"FAIL: {fails}",
        f"NEEDS MANUAL REVIEW: {manual}",
    ]
    if alerts:
        summary.append(f"LICENSE EXPIRATION ALERTS: {alerts}")

    for i, line in enumerate(summary):
        ws.cell(row=summary_start + 1 + i, column=1, value=line)

    if errors or no_admin:
        err_start = ws.max_row + 2
        ws.cell(row=err_start, column=1, value="STATES WITH ISSUES")
        ws.cell(row=err_start, column=1).font = Font(bold=True, color="9C0006")
        for i, err in enumerate(errors):
            ws.cell(row=err_start + 1 + i, column=1, value=f"ERROR: {err}")
        for i, fac in enumerate(no_admin):
            ws.cell(row=err_start + 1 + len(errors) + i, column=1,
                    value=f"NO ADMIN: {fac['state']} - {fac['facility']}")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(str(path))


if __name__ == "__main__":
    run_verification()
